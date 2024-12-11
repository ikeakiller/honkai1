import os
import platform
import csv
import webbrowser
import pyfiglet
import xlrd
import pystyle
from openpyxl import load_workbook
from colorama import init, Fore, Style
from pystyle import *

init()  # Initialize colorama

def search_in_files(search_query):
    results = []
    for file in os.listdir("bd"):
        if file.endswith((".csv", ".xls", ".xlsx", ".txt")):  # Added .xlsx to the check
            if file.endswith(".csv"):
                try:
                    with open(os.path.join("bd", file), 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            if search_query in str(row):
                                results.append(f"                                                     Найдено в {file}: {', '.join(row)}")
                except Exception as e:
                    print(f"Ошибка при чтении {file}: {e}")
            elif file.endswith(".xls"):
                try:
                    wb = xlrd.open_workbook(os.path.join("bd", file))
                    for sheet in wb.sheets():
                        for row in range(sheet.nrows):
                            if search_query in str(sheet.row_values(row)):
                                results.append(f"                                                     Найдено в {file} ({sheet.name}): {', '.join(map(str, sheet.row_values(row)))}")
                except Exception as e:
                    print(f"Ошибка при чтении {file}: {e}")
            elif file.endswith(".xlsx"):
                try:
                    wb = load_workbook(os.path.join("bd", file))
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            if search_query in str(row):
                                results.append(f"                                                     Найдено в {file} ({sheet}): {', '.join(map(str, row))}")
                except Exception as e:
                    print(f"Ошибка при чтении {file}: {e}")
            elif file.endswith(".txt"):
                try:
                    with open(os.path.join("bd", file), 'r', encoding='utf-8', errors='ignore') as f:
                        for line in f:
                            if search_query in line:
                                results.append(f"                                                     Найдено в {file}: {line.strip()}")
                except Exception as e:
                    print(f"Ошибка при чтении {file}: {e}")
    return results

search_query = input(Fore.RED + "                                                                                                Введите номер : " + Style.RESET_ALL)
results = search_in_files(search_query)

if results:
    # Clear the console
    for result in results:
        print(Fore.GREEN + result + Style.RESET_ALL)
    input(Fore.RED + "                                                                          Нажмите Enter для выхода с софта..." + Style.RESET_ALL)
else:
    print(Fore.RED + "                                                                          В базах данных не найдено результатов." + Style.RESET_ALL)
    input(Fore.WHITE + "                                                                          Нажмите Enter для выхода с софта..." + Style.RESET_ALL)